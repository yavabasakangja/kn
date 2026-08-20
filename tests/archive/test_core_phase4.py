#!/usr/bin/env python3
"""test_core_phase4.py — POC INTI **FASE 4** (KPI Saya · Dasbor Manajer · Rapor Desainer).

Menguji tiga fitur lanjutan PS-18 lewat **API NYATA** (bukan tiruan), dengan penekanan
pada hal yang paling mudah salah dan paling mahal bila salah:

  A. **KPI Saya (privasi)** — `/api/rnd/reports/my-kpi` HANYA mengirim angka milik
     pengguna yang masuk. Tidak ada nama/nilai rekan di dalam respons — dibuktikan
     dengan memeriksa SELURUH badan respons, bukan hanya field yang kita harapkan.
  B. **Dasbor Manajer** — `/api/home/manager` mengirim antrean persetujuan per jenis,
     target vs kemajuan bulan, keterlambatan hari ini (4 sumber), dan cuplikan desainer.
  C. **Rapor Desainer** — ekspor CSV/Excel/PDF benar-benar berkas sah, isinya sama
     dengan angka layar, dan hanya boleh diunduh peran penilai (admin/manager).

Jalankan: cd /app && python test_core_phase4.py
"""
import asyncio
import json
import sys
from typing import Dict

import httpx

BASE = "http://localhost:8001/api"
USERS = {
    "admin": "admin@kainnusantara.id",
    "manager": "manager@kainnusantara.id",
    "sales": "sales@kainnusantara.id",
    "warehouse": "warehouse@kainnusantara.id",
}
PASSWORD = "demo12345"
RESULTS: list = []


def check(name: str, ok: bool, detail: str = "") -> bool:
    RESULTS.append((name, bool(ok), detail))
    print(f"  {'PASS' if ok else 'FAIL'} · {name}" + (f" → {detail}" if detail else ""))
    return bool(ok)


async def login(cl: httpx.AsyncClient, role: str) -> str:
    r = await cl.post(f"{BASE}/auth/login",
                      json={"email": USERS[role], "password": PASSWORD})
    r.raise_for_status()
    # PENTING: backend mengutamakan cookie sesi (HttpOnly) di atas header Bearer
    # (`dependencies.extract_token`). Kalau cookie dibiarkan, login peran terakhir
    # akan "menimpa" semua permintaan berikutnya dan uji RBAC jadi salah baca.
    cl.cookies.clear()
    return r.json()["token"]


def hdr(token: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


# ═════════════════════════════════════════════════════════════════════════════
#  A. KPI SAYA — privasi lebih penting daripada kelengkapan
# ═════════════════════════════════════════════════════════════════════════════
async def test_my_kpi(cl: httpx.AsyncClient, tokens: Dict[str, str]) -> None:
    print("\n── A. KPI SAYA (kartu penilaian diri sendiri + privasi) ───────────")

    # Daftar SEMUA desainer (dari layar manajer) untuk uji kebocoran.
    full = (await cl.get(f"{BASE}/rnd/reports/designer-kpi?period=all",
                         headers=hdr(tokens["admin"]))).json()
    all_names = [r["designer"] for r in full["items"]]
    check("Ada >= 2 desainer di data (uji kebocoran jadi bermakna)",
          len(all_names) >= 2, ", ".join(all_names))

    mine = (await cl.get(f"{BASE}/rnd/reports/my-kpi?period=all",
                         headers=hdr(tokens["manager"]))).json()
    me = mine.get("me") or {}
    check("Manajer melihat kartu penilaiannya sendiri",
          bool(me) and me.get("designer") == mine.get("designer"),
          f"{mine.get('designer')} · grade {me.get('grade_letter')} "
          f"({me.get('grade_score')}) · peringkat {mine.get('rank')}/"
          f"{mine.get('total_designers')}")

    # KEBOCORAN: nama desainer lain tidak boleh muncul di mana pun dalam respons.
    blob = json.dumps(mine, ensure_ascii=False)
    others = [n for n in all_names if n != mine.get("designer") and n in blob]
    check("TIDAK ada nama/nilai rekan di dalam respons (disaring di server)",
          not others, f"bocor={others or 'tidak ada'}")
    check("Hanya agregat tim yang dikirim (rata-rata & jumlah, tanpa daftar)",
          isinstance(mine.get("team"), dict) and "items" not in mine
          and "leaderboard" not in mine,
          f"team={mine.get('team')}")

    check("Riwayat round yang dikirim semuanya MILIK sendiri",
          bool(mine.get("rounds")) and all(
              str(r.get("number")) for r in mine["rounds"]),
          f"{len(mine.get('rounds') or [])} round · "
          f"{len(mine.get('overdue') or [])} nunggak lewat tenggat")

    # Bobot penilaian ikut dikirim → yang dinilai tahu aturannya (transparan).
    w = mine.get("weights") or {}
    check("Aturan penilaian ikut dikirim (transparan ke yang dinilai)",
          w.get("on_time") is not None and w.get("escalate_admin_days") is not None,
          f"on_time={w.get('on_time')} score={w.get('score')} acc={w.get('acc')}")

    # Peran lain: boleh membuka miliknya sendiri, isinya kosong (bukan 403).
    for role in ("sales", "warehouse"):
        res = await cl.get(f"{BASE}/rnd/reports/my-kpi?period=all",
                           headers=hdr(tokens[role]))
        body = res.json() if res.status_code == 200 else {}
        leak = [n for n in all_names if n in json.dumps(body, ensure_ascii=False)]
        check(f"{role}: boleh buka KPI-nya sendiri & tidak melihat data rekan",
              res.status_code == 200 and not body.get("me") and not leak,
              f"HTTP {res.status_code} · me={bool(body.get('me'))} · bocor={leak}")

    # Filter periode ikut jalan di kartu pribadi.
    m30 = (await cl.get(f"{BASE}/rnd/reports/my-kpi?period=30d",
                        headers=hdr(tokens["manager"]))).json()
    check("Filter periode berlaku juga di KPI Saya",
          m30.get("period") == "30d" and bool(m30.get("from_date")),
          f"{m30.get('period_label')} sejak {m30.get('from_date')}")


# ═════════════════════════════════════════════════════════════════════════════
#  B. DASBOR MANAJER
# ═════════════════════════════════════════════════════════════════════════════
async def test_manager_home(cl: httpx.AsyncClient, tokens: Dict[str, str]) -> None:
    print("\n── B. DASBOR MANAJER (persetujuan · target tim · terlambat) ───────")
    res = await cl.get(f"{BASE}/home/manager", headers=hdr(tokens["manager"]))
    check("Endpoint /home/manager terbuka untuk manajer", res.status_code == 200,
          f"HTTP {res.status_code}")
    d = res.json()

    ap = d.get("approvals") or {}
    check("Antrean persetujuan dirinci per jenis + tujuan klik",
          isinstance(ap.get("all_items"), list) and len(ap["all_items"]) >= 4
          and all(r.get("view") for r in ap["all_items"]),
          f"total={ap.get('total')} · " + " · ".join(
              f"{r['label']}={r['count']}" for r in ap.get("all_items", [])))

    tgt = d.get("target") or {}
    check("Target tim dibandingkan dengan KEMAJUAN BULAN (bukan angka telanjang)",
          tgt.get("amount") is not None and tgt.get("achievement_pct") is not None
          and tgt.get("month_progress_pct") is not None,
          f"target={tgt.get('amount')} capaian={tgt.get('achievement_pct')}% "
          f"bulan berjalan={tgt.get('month_progress_pct')}% "
          f"(hari {tgt.get('day')}/{tgt.get('days_in_month')})")

    late = d.get("late_today") or {}
    keys = {r["key"] for r in late.get("all_rows", [])}
    check("Keterlambatan hari ini mencakup 4 sumber (piutang · R&D · gudang · produksi)",
          keys == {"ar", "rnd", "wms", "production"},
          f"{late.get('total_items')} hal · " + " · ".join(
              f"{r['key']}={r['count'] if r['count'] is not None else r['amount']}"
              for r in late.get("all_rows", [])))
    check("Angka round R&D di dasbor SAMA dengan papan eskalasi",
          late.get("rnd_overdue") == (await cl.get(
              f"{BASE}/rnd/sla/board", headers=hdr(tokens["manager"]))).json()["count"],
          f"dasbor={late.get('rnd_overdue')} · naik ke admin="
          f"{late.get('rnd_escalated_admin')}")

    team = d.get("team") or []
    check("Tiap anggota tim punya target & capaiannya sendiri",
          bool(team) and all("target_collection" in t and "achievement_pct" in t
                             for t in team),
          ", ".join(f"{t['sales_name']}={t['achievement_pct']}%" for t in team[:4]))

    des = d.get("designers") or {}
    check("Cuplikan kinerja desainer ikut tampil untuk manajer",
          des.get("count", 0) >= 1 and bool(des.get("top")),
          f"{des.get('count')} desainer · terbaik="
          f"{(des.get('summary') or {}).get('best_designer')}")

    # RBAC: sales & gudang tidak boleh membuka dasbor manajer.
    for role in ("sales", "warehouse"):
        r = await cl.get(f"{BASE}/home/manager", headers=hdr(tokens[role]))
        check(f"{role} DITOLAK membuka dasbor manajer", r.status_code == 403,
              f"HTTP {r.status_code}")
    r_admin = await cl.get(f"{BASE}/home/manager", headers=hdr(tokens["admin"]))
    check("Admin tetap boleh membuka dasbor manajer", r_admin.status_code == 200,
          f"HTTP {r_admin.status_code}")


# ═════════════════════════════════════════════════════════════════════════════
#  C. RAPOR DESAINER (ekspor)
# ═════════════════════════════════════════════════════════════════════════════
async def test_export(cl: httpx.AsyncClient, tokens: Dict[str, str]) -> None:
    print("\n── C. RAPOR DESAINER (unduh CSV · Excel · PDF) ────────────────────")
    rep = (await cl.get(f"{BASE}/rnd/reports/designer-kpi?period=all",
                        headers=hdr(tokens["admin"]))).json()
    rows = rep["count"]
    names = [r["designer"] for r in rep["items"]]

    magic = {"csv": b"\xef\xbb\xbf", "xlsx": b"PK", "pdf": b"%PDF"}
    for fmt in ("csv", "xlsx", "pdf"):
        r = await cl.get(f"{BASE}/rnd/reports/designer-kpi/export"
                         f"?period=all&format={fmt}", headers=hdr(tokens["admin"]))
        body = r.content
        disp = r.headers.get("content-disposition", "")
        check(f"Unduh {fmt.upper()} menghasilkan berkas sah + nama berkas jelas",
              r.status_code == 200 and body.startswith(magic[fmt])
              and "attachment; filename=" in disp and len(body) > 500,
              f"HTTP {r.status_code} · {len(body)} byte · {disp[:60]}")
        if fmt == "csv":
            text = body.decode("utf-8-sig")
            check("CSV memuat SEMUA desainer + judul & rumus penilaian",
                  all(n in text for n in names) and "Laporan KPI Desainer" in text
                  and "Nilai =" in text,
                  f"{len(text.splitlines())} baris untuk {rows} desainer")
        if fmt == "xlsx":
            check("Excel bisa dibuka kembali & jumlah barisnya cocok",
                  _xlsx_rows(body) == rows, f"{_xlsx_rows(body)} baris data (harap {rows})")

    # Format asal-asalan ditolak dengan pesan jelas, bukan 500.
    bad = await cl.get(f"{BASE}/rnd/reports/designer-kpi/export?format=docx",
                       headers=hdr(tokens["admin"]))
    check("Format tak dikenal ditolak dengan pesan jelas (bukan error 500)",
          bad.status_code == 400 and "csv" in bad.text.lower(),
          f"HTTP {bad.status_code} · {bad.text[:80]}")

    # Periode ikut menentukan isi berkas.
    r30 = await cl.get(f"{BASE}/rnd/reports/designer-kpi/export"
                       f"?period=30d&format=csv", headers=hdr(tokens["admin"]))
    check("Berkas mengikuti periode yang dipilih di layar",
          "30 hari terakhir" in r30.content.decode("utf-8-sig"),
          r30.headers.get("content-disposition", "")[:60])

    # RBAC: rapor orang hanya untuk peran penilai.
    for role, expect in (("manager", 200), ("sales", 403), ("warehouse", 403)):
        r = await cl.get(f"{BASE}/rnd/reports/designer-kpi/export?format=csv",
                         headers=hdr(tokens[role]))
        check(f"{role}: unduh rapor desainer → HTTP {expect}",
              r.status_code == expect, f"HTTP {r.status_code}")


def _xlsx_rows(data: bytes) -> int:
    from io import BytesIO

    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data)).active
    head = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "#":
            head = i
            break
    if head is None:
        return -1
    return sum(1 for row in ws.iter_rows(min_row=head + 1, values_only=True)
               if row and row[0] is not None)


async def main() -> int:
    print("=" * 78)
    print("POC FASE 4 — KPI Saya · Dasbor Manajer · Rapor Desainer")
    print("=" * 78)
    async with httpx.AsyncClient(timeout=60) as cl:
        tokens = {role: await login(cl, role) for role in USERS}
        print(f"  login OK untuk {len(tokens)} peran")
        await test_my_kpi(cl, tokens)
        await test_manager_home(cl, tokens)
        await test_export(cl, tokens)

    passed = sum(1 for _, ok, _ in RESULTS if ok)
    print("\n" + "=" * 78)
    print(f"HASIL: {passed}/{len(RESULTS)} lulus")
    for name, ok, detail in RESULTS:
        if not ok:
            print(f"  FAIL · {name} → {detail}")
    print("=" * 78)
    return 0 if passed == len(RESULTS) else 1


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
