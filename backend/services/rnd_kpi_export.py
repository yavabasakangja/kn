"""PS-18 — **EKSPOR LAPORAN KPI DESAINER** (CSV · Excel · PDF).

Untuk apa: rapat bulanan tidak dibuka di depan layar aplikasi. Pemilik butuh berkas
yang bisa dicetak, ditempel di notulen, atau dikirim ke pemegang saham. Karena itu
ketiga format dibuat dari **SATU sumber angka** (`rnd_kpi_service.designer_kpi`) —
tidak ada perhitungan kedua yang bisa berbeda dari layar.

Pilihan teknologi mengikuti yang SUDAH dipakai repo ini (tidak menambah dependensi):
`csv` (bawaan) · `openpyxl` (dipakai impor barang supplier) · `reportlab`
(dipakai slip gaji `hr_payroll_pdf.py`).
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Tuple

from core_utils import now_iso, rupiah

# Kolom laporan — SATU definisi dipakai CSV, Excel, dan PDF supaya isi ketiga
# berkas tidak mungkin berbeda. `w` = lebar kolom PDF (mm), `x` = lebar Excel.
COLUMNS: Tuple[Tuple[str, str, float, float], ...] = (
    ("rank", "#", 8, 5),
    ("designer", "Desainer", 42, 24),
    ("samples", "Permintaan", 20, 12),
    ("rounds", "Round", 14, 8),
    ("submitted", "Disetor", 16, 9),
    ("assessed", "Dinilai", 16, 9),
    ("acc", "ACC", 12, 7),
    ("revisi", "Revisi", 14, 8),
    ("tolak", "Tolak", 14, 8),
    ("late_submitted", "Setor telat", 20, 12),
    ("overdue_now", "Nunggak", 17, 10),
    ("on_time_pct", "Tepat waktu %", 23, 14),
    ("rework_pct", "Diulang %", 19, 12),
    ("acc_rate", "ACC-rate %", 20, 12),
    ("avg_score", "Rata skor", 19, 11),
    ("avg_days", "Rata hari", 19, 11),
    ("grade_base", "Nilai dasar", 20, 12),
    ("grade_penalty", "Penalti", 16, 10),
    ("grade_score", "Nilai akhir", 20, 12),
    ("grade_letter", "Grade", 14, 8),
    ("cost_total", "Biaya sample", 28, 16),
)

MONEY_KEYS = {"cost_total"}


def _cell(row: Dict[str, Any], key: str) -> Any:
    val = row.get(key)
    if val is None:
        return "—"
    return val


def _txt(row: Dict[str, Any], key: str) -> str:
    val = _cell(row, key)
    if key in MONEY_KEYS and val != "—":
        return rupiah(val)
    return str(val)


def title_of(rep: Dict[str, Any]) -> str:
    return f"Laporan KPI Desainer — {rep.get('period_label', '')}"


def subtitle_of(rep: Dict[str, Any], entity_name: str = "") -> str:
    rentang = (f"{rep.get('from_date')} s/d {rep.get('to_date')}"
               if rep.get("from_date") else f"seluruh data s/d {rep.get('to_date')}")
    head = f"{entity_name} · " if entity_name else ""
    return f"{head}Periode {rentang} · dibuat {now_iso()[:16].replace('T', ' ')}"


def formula_of(rep: Dict[str, Any]) -> str:
    w = rep.get("weights") or {}
    return (f"Nilai = tepat waktu {w.get('on_time', 0)}% + skor penilaian "
            f"{w.get('score', 0)}% + sekali-jadi (ACC) {w.get('acc', 0)}%, dikurangi "
            f"penalti diulang ({w.get('penalty_rework', 0)}x) dan penalti terlambat "
            f"({w.get('penalty_overdue', 0)}x). Eskalasi ke admin setelah "
            f"{w.get('escalate_admin_days', 3)} hari terlambat.")


def summary_lines(rep: Dict[str, Any]) -> List[str]:
    s = rep.get("summary") or {}
    return [
        f"Desainer aktif: {s.get('designers', 0)} · Round dikerjakan: {s.get('rounds', 0)}"
        f" · Disetor: {s.get('submitted', 0)} · Dinilai: {s.get('assessed', 0)}",
        f"Tepat waktu: {s.get('on_time_pct', '—')}% · Diulang: {s.get('rework_pct', '—')}%"
        f" · Rata skor: {s.get('avg_score', '—')} · Rata nilai: {s.get('avg_grade', '—')}",
        f"Masih nunggak lewat tenggat: {s.get('overdue_now', 0)} round"
        f" ({s.get('overdue_critical', 0)} sudah naik ke admin)"
        f" · Biaya sample: {rupiah(s.get('cost_total') or 0)}",
        f"Terbaik periode ini: {s.get('best_designer') or '—'}"
        f" ({s.get('best_grade') or '—'})",
    ]


def filename(rep: Dict[str, Any], ext: str) -> str:
    per = str(rep.get("period") or "all")
    return f"kpi-desainer-{per}-{now_iso()[:10]}.{ext}"


# ─── CSV ─────────────────────────────────────────────────────────────────────
def csv_bytes(rep: Dict[str, Any], entity_name: str = "") -> bytes:
    buf = io.StringIO()
    wr = csv.writer(buf)
    wr.writerow([title_of(rep)])
    wr.writerow([subtitle_of(rep, entity_name)])
    wr.writerow([formula_of(rep)])
    for line in summary_lines(rep):
        wr.writerow([line])
    wr.writerow([])
    wr.writerow([label for _, label, _, _ in COLUMNS])
    for row in rep.get("items") or []:
        wr.writerow([_cell(row, key) for key, _, _, _ in COLUMNS])
    # BOM supaya Excel di Windows membaca huruf beraksen dengan benar.
    return buf.getvalue().encode("utf-8-sig")


# ─── EXCEL (openpyxl) ────────────────────────────────────────────────────────
def xlsx_bytes(rep: Dict[str, Any], entity_name: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Desainer"
    navy = "FF0058CC"
    thin = Side(style="thin", color="FFE1E4E8")

    ws["A1"] = title_of(rep)
    ws["A1"].font = Font(bold=True, size=14, color=navy)
    ws["A2"] = subtitle_of(rep, entity_name)
    ws["A3"] = formula_of(rep)
    ws["A3"].alignment = Alignment(wrap_text=False)
    line_no = 4
    for line in summary_lines(rep):
        ws.cell(row=line_no, column=1, value=line)
        line_no += 1
    head_row = line_no + 1

    for idx, (_, label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=head_row, column=idx, value=label)
        cell.font = Font(bold=True, size=10, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
        ws.column_dimensions[cell.column_letter].width = width

    for r_i, row in enumerate(rep.get("items") or [], start=head_row + 1):
        for c_i, (key, _, _, _) in enumerate(COLUMNS, start=1):
            val = row.get(key)
            cell = ws.cell(row=r_i, column=c_i,
                           value="—" if val is None else val)
            cell.border = Border(bottom=thin)
            if key in MONEY_KEYS and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
            if key == "designer":
                cell.font = Font(bold=True, size=10)
            elif key == "grade_letter":
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── PDF (reportlab — pola sama slip gaji) ───────────────────────────────────
def pdf_bytes(rep: Dict[str, Any], entity_name: str = "Kain Nusantara") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    NAVY = colors.HexColor("#0058CC")
    DARK = colors.HexColor("#1A1A1F")
    GRAY = colors.HexColor("#6B6B73")
    LINE = colors.HexColor("#E1E4E8")
    GREEN = colors.HexColor("#1B7F4B")
    RED = colors.HexColor("#C0392B")
    ORANGE = colors.HexColor("#B26A00")
    GRADE_COLOR = {"A": GREEN, "B": NAVY, "C": ORANGE, "D": RED}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=title_of(rep), author=entity_name)
    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=14, leading=17,
                        textColor=NAVY, alignment=TA_LEFT)
    small = ParagraphStyle("small", fontName="Helvetica", fontSize=7.5, leading=10,
                           textColor=GRAY)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=8, leading=11,
                          textColor=DARK)
    cellhead = ParagraphStyle("ch", fontName="Helvetica-Bold", fontSize=6.6,
                              leading=8, textColor=colors.white)

    flow: List[Any] = [Paragraph(title_of(rep), h1),
                       Paragraph(subtitle_of(rep, entity_name), small),
                       Spacer(1, 3 * mm)]
    for line in summary_lines(rep):
        flow.append(Paragraph(line, body))
    flow += [Spacer(1, 2 * mm), Paragraph(formula_of(rep), small), Spacer(1, 3 * mm)]

    head = [Paragraph(label, cellhead) for _, label, _, _ in COLUMNS]
    data: List[List[Any]] = [head]
    for row in rep.get("items") or []:
        data.append([_txt(row, key) for key, _, _, _ in COLUMNS])

    widths = [w * mm for _, _, w, _ in COLUMNS]
    table = Table(data, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
    ]
    grade_col = [i for i, (k, _, _, _) in enumerate(COLUMNS) if k == "grade_letter"][0]
    for r_i, row in enumerate(rep.get("items") or [], start=1):
        if r_i % 2 == 0:
            style.append(("BACKGROUND", (0, r_i), (-1, r_i),
                          colors.HexColor("#FAFBFC")))
        tone = GRADE_COLOR.get(str(row.get("grade_letter") or ""), GRAY)
        style += [("TEXTCOLOR", (grade_col, r_i), (grade_col, r_i), tone),
                  ("FONTNAME", (grade_col, r_i), (grade_col, r_i), "Helvetica-Bold")]
    table.setStyle(TableStyle(style))
    flow.append(table)

    if not (rep.get("items") or []):
        flow.append(Paragraph(
            "Belum ada round sample yang disetor pada periode ini, sehingga belum ada "
            "kinerja desainer yang bisa dinilai.", body))
    flow += [Spacer(1, 4 * mm), Paragraph(
        "Angka pada laporan ini terbentuk otomatis dari jejak round sample "
        "(lampiran + catatan wajib) — tidak ada nilai yang diisi manual.", small)]

    doc.build(flow)
    return buf.getvalue()


FORMATS: Dict[str, Tuple[str, str]] = {
    "csv": ("csv", "text/csv; charset=utf-8"),
    "xlsx": ("xlsx",
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "pdf": ("pdf", "application/pdf"),
}


# ═════════════════════════════════════════════════════════════════════════════
#  RAPOR PER-DESAINER (1 halaman PDF) — lampiran evaluasi (PS-18 lanjutan)
#
#  Pemilik ingin memberi UMPAN BALIK ke satu orang saat evaluasi, bukan menyodorkan
#  tabel seluruh tim. Rapor ini seperti "kartu rapor": grade besar di atas, metrik
#  kunci, pembanding tim yang AGREGAT (tanpa nama rekan), lalu riwayat round terbaru
#  sebagai bukti konkret. Sumber angka SAMA (`my_kpi`) → tidak ada hitungan kedua.
# ═════════════════════════════════════════════════════════════════════════════
def _slug(text: str) -> str:
    keep = [c.lower() if c.isalnum() else "-" for c in str(text or "").strip()]
    out = "".join(keep)
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-") or "desainer"


def designer_report_filename(mine: Dict[str, Any], ext: str = "pdf") -> str:
    per = str(mine.get("period") or "all")
    return f"rapor-desainer-{_slug(mine.get('designer'))}-{per}-{now_iso()[:10]}.{ext}"


def _band_meaning(mine: Dict[str, Any], letter: str) -> str:
    for b in mine.get("grade_bands") or []:
        if str(b.get("letter")) == str(letter):
            return str(b.get("meaning") or "")
    return ""


def designer_report_pdf(mine: Dict[str, Any], entity_name: str = "Kain Nusantara",
                        note: str = "") -> bytes:
    """Rapor kinerja SATU desainer dalam 1 halaman A4 (portrait).

    `note` = catatan evaluasi bebas dari penilai (opsional) — ditulis manager sebelum
    unduh, lalu tampil sebagai kotak "Catatan Evaluasi" agar rapor bisa langsung
    ditandatangani/ditempel di berkas evaluasi.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    NAVY = colors.HexColor("#0058CC")
    DARK = colors.HexColor("#1A1A1F")
    GRAY = colors.HexColor("#6B6B73")
    LINE = colors.HexColor("#E1E4E8")
    SOFT = colors.HexColor("#FAFBFC")
    GREEN = colors.HexColor("#1B7F4B")
    RED = colors.HexColor("#C0392B")
    ORANGE = colors.HexColor("#B26A00")
    GRADE_COLOR = {"A": GREEN, "B": NAVY, "C": ORANGE, "D": RED}

    me = mine.get("me") or {}
    name = mine.get("designer") or "—"
    letter = str(me.get("grade_letter") or "—")
    tone = GRADE_COLOR.get(letter, GRAY)
    rentang = (f"{mine.get('from_date')} s/d {mine.get('to_date')}"
               if mine.get("from_date") else f"seluruh data s/d {mine.get('to_date')}")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title=f"Rapor Desainer — {name}", author=entity_name)
    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=15, leading=18,
                        textColor=NAVY, alignment=TA_LEFT)
    small = ParagraphStyle("small", fontName="Helvetica", fontSize=8, leading=11,
                           textColor=GRAY)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=9, leading=13,
                          textColor=DARK)
    label_st = ParagraphStyle("lab", fontName="Helvetica", fontSize=7, leading=9,
                              textColor=GRAY, alignment=TA_CENTER)
    val_st = ParagraphStyle("val", fontName="Helvetica-Bold", fontSize=12, leading=14,
                            textColor=DARK, alignment=TA_CENTER)
    grade_big = ParagraphStyle("gb", fontName="Helvetica-Bold", fontSize=34, leading=36,
                               textColor=colors.white, alignment=TA_CENTER)
    grade_sub = ParagraphStyle("gs", fontName="Helvetica", fontSize=8, leading=10,
                               textColor=colors.white, alignment=TA_CENTER)
    cellhead = ParagraphStyle("ch", fontName="Helvetica-Bold", fontSize=7.2,
                              leading=9, textColor=colors.white)

    flow: List[Any] = []

    # ── Kepala: judul (kiri) + kartu grade besar (kanan) ──────────────────────
    rank_txt = (f"Peringkat {mine.get('rank')} dari {mine.get('total_designers')}"
                if mine.get("rank") else "Belum masuk peringkat")
    left = [Paragraph("Rapor Kinerja Desainer", h1),
            Paragraph(f"<b>{name}</b>", body),
            Paragraph(f"{entity_name} · Periode {mine.get('period_label', '')} "
                      f"({rentang})", small),
            Paragraph(f"Dibuat {now_iso()[:16].replace('T', ' ')} · {rank_txt}", small)]
    score_txt = ("—" if me.get("grade_score") is None else str(me.get("grade_score")))
    grade_box = Table(
        [[Paragraph(letter, grade_big)],
         [Paragraph(f"Nilai {score_txt} / 100", grade_sub)],
         [Paragraph(_band_meaning(mine, letter) or "kinerja", grade_sub)]],
        colWidths=[42 * mm])
    grade_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), tone),
        ("BOX", (0, 0), (-1, -1), 0, tone),
        ("TOPPADDING", (0, 0), (-1, 0), 6), ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, 0), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
    ]))
    header = Table([[left, grade_box]], colWidths=[122 * mm, 46 * mm])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    flow += [header, Spacer(1, 5 * mm)]

    if not me:
        flow.append(Paragraph(
            "Desainer ini belum punya round sample yang disetor pada periode ini, "
            "sehingga belum ada kinerja yang bisa dinilai.", body))
        doc.build(flow)
        return buf.getvalue()

    # ── Kartu metrik kunci (grid 4 kolom) ─────────────────────────────────────
    def _num(v, suffix=""):
        return "—" if v is None else f"{v}{suffix}"

    metrics = [
        ("Round dikerjakan", _num(me.get("rounds"))),
        ("Disetor", _num(me.get("submitted"))),
        ("Dinilai", _num(me.get("assessed"))),
        ("Tepat waktu", _num(me.get("on_time_pct"), "%")),
        ("ACC (sekali jadi)", _num(me.get("acc"))),
        ("Revisi", _num(me.get("revisi"))),
        ("Tolak", _num(me.get("tolak"))),
        ("Kerja diulang", _num(me.get("rework_pct"), "%")),
        ("Rata skor mutu", _num(me.get("avg_score"))),
        ("Rata hari selesai", _num(me.get("avg_days"))),
        ("Setor terlambat", _num(me.get("late_submitted"))),
        ("Nunggak lewat tenggat", _num(me.get("overdue_now"))),
    ]
    cards = []
    row: List[Any] = []
    for lab, val in metrics:
        row.append(Table([[Paragraph(lab, label_st)], [Paragraph(str(val), val_st)]],
                         colWidths=[40 * mm]))
        if len(row) == 4:
            cards.append(row)
            row = []
    if row:
        while len(row) < 4:
            row.append("")
        cards.append(row)
    grid = Table(cards, colWidths=[42 * mm] * 4)
    gstyle = [("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
              ("BOX", (0, 0), (-1, -1), 0.5, LINE),
              ("INNERGRID", (0, 0), (-1, -1), 0.5, LINE),
              ("TOPPADDING", (0, 0), (-1, -1), 5),
              ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]
    grid.setStyle(TableStyle(gstyle))
    flow += [grid, Spacer(1, 4 * mm)]

    # ── Pembanding tim (AGREGAT — tanpa nama rekan) + biaya ────────────────────
    team = mine.get("team") or {}
    flow += [Paragraph(
        f"<b>Pembanding tim (rata-rata {team.get('designers', 0)} desainer):</b> "
        f"tepat waktu {_num(team.get('on_time_pct'), '%')} · diulang "
        f"{_num(team.get('rework_pct'), '%')} · rata skor {_num(team.get('avg_score'))} "
        f"· rata nilai tim {_num(team.get('avg_grade'))}. "
        f"Biaya sample orang ini: {rupiah(me.get('cost_total') or 0)}.", body),
        Spacer(1, 4 * mm)]

    # ── Riwayat round terbaru (bukti konkret) ─────────────────────────────────
    RH_COLS = (("number", "Permintaan", 28), ("round_no", "Round", 12),
               ("result", "Hasil", 18), ("score", "Skor", 14),
               ("due_date", "Tenggat", 24), ("submitted_at", "Disetor", 24),
               ("on_time", "Setor tepat?", 18), ("days_late", "Nunggak (hr)", 20))
    rounds = mine.get("rounds") or []
    flow.append(Paragraph("Riwayat round terbaru (bukti penilaian)", ParagraphStyle(
        "rh", fontName="Helvetica-Bold", fontSize=10, leading=12, textColor=NAVY)))
    flow.append(Spacer(1, 1.5 * mm))
    if rounds:
        head = [Paragraph(lab, cellhead) for _, lab, _ in RH_COLS]
        data = [head]
        for r in rounds[:12]:
            cells = []
            for key, _, _ in RH_COLS:
                v = r.get(key)
                if key == "on_time":
                    v = "—" if v is None else ("Ya" if v else "Telat")
                elif key == "score":
                    v = "—" if v is None else v
                elif key == "days_late":
                    v = 0 if v is None else v
                cells.append(str(v if v not in (None, "") else "—"))
            data.append(cells)
        widths = [w * mm for _, _, w in RH_COLS]
        rt = Table(data, colWidths=widths, repeatRows=1)
        rstyle = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("GRID", (0, 0), (-1, -1), 0.4, LINE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i in range(1, len(data)):
            if i % 2 == 0:
                rstyle.append(("BACKGROUND", (0, i), (-1, i), SOFT))
        rt.setStyle(TableStyle(rstyle))
        flow.append(rt)
        if len(rounds) > 12:
            flow.append(Spacer(1, 1 * mm))
            flow.append(Paragraph(
                f"Menampilkan 12 round terbaru dari {len(rounds)} round pada periode ini. "
                "Kolom \u201cSetor tepat?\u201d = apakah hasil disetor sebelum tenggat; "
                "\u201cNunggak (hr)\u201d = hari terlambat untuk round yang MASIH berjalan.",
                small))
        else:
            flow.append(Spacer(1, 1 * mm))
            flow.append(Paragraph(
                "Kolom \u201cSetor tepat?\u201d = apakah hasil disetor sebelum tenggat; "
                "\u201cNunggak (hr)\u201d = hari terlambat untuk round yang MASIH berjalan.",
                small))
    else:
        flow.append(Paragraph("Belum ada round pada periode ini.", body))

    # ── Catatan evaluasi (opsional, ditulis penilai sebelum unduh) ─────────────
    import html as _html
    note_txt = (note or "").strip()
    safe_note = (_html.escape(note_txt).replace("\n", "<br/>") if note_txt
                 else "&nbsp;<br/>&nbsp;<br/>&nbsp;")
    note_tbl = Table(
        [[Paragraph("Catatan Evaluasi", ParagraphStyle(
            "et", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=NAVY))],
         [Paragraph(safe_note, body)]],
        colWidths=[168 * mm])
    note_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, LINE),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, LINE),
        ("BACKGROUND", (0, 0), (-1, 0), SOFT),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    flow += [Spacer(1, 4 * mm), note_tbl]

    flow += [Spacer(1, 4 * mm), Paragraph(formula_of(mine), small), Spacer(1, 1 * mm),
             Paragraph("Angka pada rapor ini terbentuk otomatis dari jejak round sample "
                       "(lampiran + catatan wajib) — tidak ada nilai yang diisi manual.",
                       small)]
    doc.build(flow)
    return buf.getvalue()


def render_designer_report(mine: Dict[str, Any], fmt: str = "pdf",
                           entity_name: str = "", note: str = "") -> Tuple[bytes, str, str]:
    """Return (bytes, filename, media_type) untuk RAPOR SATU desainer. Saat ini: pdf."""
    key = (fmt or "pdf").strip().lower()
    if key != "pdf":
        raise ValueError("Rapor per-desainer saat ini hanya tersedia dalam format PDF.")
    data = designer_report_pdf(mine, entity_name or "Kain Nusantara", note=note)
    return data, designer_report_filename(mine, "pdf"), FORMATS["pdf"][1]


def render(rep: Dict[str, Any], fmt: str, entity_name: str = "") -> Tuple[bytes, str, str]:
    """Return (bytes, filename, media_type). `fmt` = csv | xlsx | pdf."""
    key = (fmt or "xlsx").strip().lower()
    if key not in FORMATS:
        raise ValueError("Format ekspor harus salah satu: csv, xlsx, pdf.")
    ext, media = FORMATS[key]
    if key == "csv":
        data = csv_bytes(rep, entity_name)
    elif key == "xlsx":
        data = xlsx_bytes(rep, entity_name)
    else:
        data = pdf_bytes(rep, entity_name or "Kain Nusantara")
    return data, filename(rep, ext), media
